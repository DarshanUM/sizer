import os
import openpyxl
import json
import requests
from copy import deepcopy
from sys import argv
from collections import defaultdict

from openpyxl.styles import PatternFill, colors
from mailer import Mailer
from mailer import Message
from wl_template import VDI, VSI, ROBO, DB, ORACLE, RAW, MIXED, VDI_INFRA, VEEAM, SPLUNK, RDSH, AIML

WL_TYPE_MAPPER = {'VDI': VDI,
                  'VSI': VSI,
                  'ROBO': ROBO,
                  'DB': DB,
                  'ORACLE': ORACLE,
                  'RAW': RAW,
                  'INFRA': VDI_INFRA,
                  'VEEAM': VEEAM,
                  'SPLUNK': SPLUNK,
                  'RDSH': RDSH,
                  'AIML':AIML}

MAX_LIMIT = {"VDI": 30000,
             "RDSH": 30000,
             "VSI": 10000,
             "ROBO": 2000,
             "DB": 2000,
             "ORACLE": 2000}

CONVERSION_DICT = {'Windows 7': 'win_7',
                   'Windows 10': 'win_10',
                   'Task': 'Task Worker',
                   'Power': 'Power User',
                   'Knowledge': 'Knowledge Worker',
                   'Custom': 'Custom User',
                   'Pooled': 'Pooled Desktops',
                   'Citrix': 'citrix',
                   'Horizon': 'horizon',
                   'Persistent': 'Persistent Desktops',
                   'cto_hybrid': ['HX220C', 'HX240C', 'HX240C [LFF]', 'HX240C [12TB LFF]', 'HX-E-220'],
                   'cto_flash': ['HXAF-220', 'HXAF-240', 'HXAF-220 [NVME]', 'HXAF-E-220'],
                   'bundle_hybrid': ['HX-SP-220', 'HX-SP-240', 'HX-SP-240 [LFF]'],
                   'bundle_flash': ['HXAF-SP-220', 'HXAF-SP-240'],
                   'Conservative': 0,
                   'Standard': 1,
                   'Aggressive': 2,
                   'HX + SPLUNK': 'hx+splunk',
                   'HX + SPLUNK Smartstore': 'hx+splunk_smartstore'}

PART_CONVERSION_DICT = {'16': '16GiB DDR4',
                        '32': '32GiB DDR4',
                        '64': '64GiB DDR4',
                        '128': '128GiB DDR4',
                        '960': '960GB [SSD]',
                        '1.2': '1.2TB [HDD]',
                        '1.8': '1.8TB [HDD]',
                        '3.8': '3.8TB [SSD]'}


RESULT_SET = {"Max #", "HX Count", "HX Node", "HX Desc", "CO Count", "CO Node", "CO Desc", "Accessory Count",
              "Accessories", "Utilization"}

if argv[1] == 'Production':
    from production_test_conf import TEMPLATE, SCENARIO_ID, RESULT_FILE_NAME
elif argv[1] == 'Master':
    from master_test_conf import TEMPLATE, SCENARIO_ID, RESULT_FILE_NAME
else:
    raise Exception("Wrong Argument")


def initialise_test_dict():

    test_dict = {"tests": 0,
                 "regressions": 0,
                 "expected errors": 0,
                 "success": 0}

    for sheet_name in workload_tests:
        error_count[sheet_name] = deepcopy(test_dict)


def get_cell_map(sheet):

    cell_map = dict()

    for col in range(1, sheet.max_column + 1):
        cell_map[sheet.cell(row=1, column=col).value] = dict()
        cell_map[sheet.cell(row=1, column=col).value]["data"] = None
        cell_map[sheet.cell(row=1, column=col).value]["column"] = col

    return cell_map


def get_row_map(sheet, column_num):

    workload_set_map = list()

    previous_set_start = 2

    for row_num in range(3, sheet.max_row + 1):

        if sheet.cell(row=row_num, column=column_num).value:
            workload_set_map.append((previous_set_start, row_num - 1))
            previous_set_start = row_num

    workload_set_map.append((previous_set_start, row_num))

    return workload_set_map


def edit_settings(workload, cell_map):

    for setting_count in range(len(workload["settings_json"])):

        workload["settings_json"][setting_count]['sizer_version'] = version

        if cell_map['hypervisor']['data'] == 'ESXi':
            workload["settings_json"][setting_count]['hypervisor'] = 'esxi'
        elif cell_map['hypervisor']['data'] == 'Hyper-V':
            workload["settings_json"][setting_count]['hypervisor'] = 'hyperv'

        if cell_map['Bundle']['data'] == 'CTO':
            workload["settings_json"][setting_count]['bundle_only'] = False
        else:
            workload["settings_json"][setting_count]['bundle_only'] = True

        if cell_map['Node']['data'] == 'HX + CO':
            workload["settings_json"][setting_count]["heterogenous"] = True
        else:
            workload["settings_json"][setting_count]["heterogenous"] = False

        if "Threshold" in cell_map and cell_map['Threshold']["data"]:
            workload["settings_json"][setting_count]["threshold"] = CONVERSION_DICT[cell_map['Threshold']["data"]]
        else:
            workload["settings_json"][setting_count]["threshold"] = CONVERSION_DICT['Standard']

        if 'server_type' in cell_map and cell_map['server_type']['data']:
            workload["settings_json"][setting_count]["server_type"] = cell_map["server_type"]["data"]
        else:
            workload["settings_json"][setting_count]["server_type"] = 'ALL'

        if 'Disk type' in cell_map and cell_map['Disk type']['data']:
            workload["settings_json"][setting_count]["disk_option"] = cell_map["Disk type"]['data']
        else:
            workload["settings_json"][setting_count]["disk_option"] = 'ALL'

        if 'Cache type' in cell_map and cell_map['Cache type']['data']:
            workload["settings_json"][setting_count]["cache_option"] = cell_map["Cache type"]['data']
        else:
            workload["settings_json"][setting_count]["cache_option"] = 'ALL'

        if 'modular_lan' and cell_map['modular_lan']['data']:
            workload["settings_json"][setting_count]["modular_lan"] = cell_map["modular_lan"]['data']
        else:
            workload["settings_json"][setting_count]["modular_lan"] = 'ALL'

        if cell_map['Flash']['data'] == 'All Flash':

            if workload["settings_json"][setting_count]['bundle_only']:
                workload["settings_json"][setting_count]['filters']['Node_Type'] = CONVERSION_DICT['bundle_flash']
            else:
                workload["settings_json"][setting_count]['filters']['Node_Type'] = CONVERSION_DICT['cto_flash']

        else:

            if workload["settings_json"][setting_count]['bundle_only']:
                workload["settings_json"][setting_count]['filters']['Node_Type'] = CONVERSION_DICT['bundle_hybrid']
            else:
                workload["settings_json"][setting_count]['filters']['Node_Type'] = CONVERSION_DICT['cto_hybrid']

        for filter_option_key in ["RAM_Slots", "RAM_Options", "Disk_Options", "Cache_Options", "Node_Type", "CPU_Type",
                                  "Compute_Type", "GPU_Type"]:

            if filter_option_key in cell_map and cell_map[filter_option_key]["data"]:

                if filter_option_key == "RAM_Slots":

                    workload["settings_json"][setting_count]['filters'][filter_option_key] = \
                        [int(slot.strip()) for slot in str(cell_map["RAM_Slots"]["data"]).split(',')]

                elif filter_option_key == "Node_Type":

                    node_list = cell_map[filter_option_key]['data'].split(',')

                    workload['settings_json'][setting_count]['filters'][filter_option_key] = \
                        [filter_value.strip() for filter_value in node_list]

                else:

                    workload['settings_json'][setting_count]['filters'][filter_option_key] = list()

                    for filter_part in str(cell_map[filter_option_key]["data"]).split(','):

                        filter_part = filter_part.strip()

                        if filter_option_key in ["RAM_Options", "Disk_Options", "Cache_Options"]:

                            try:
                                workload["settings_json"][setting_count]['filters'][filter_option_key]. \
                                    append(PART_CONVERSION_DICT[filter_part])
                            except KeyError:
                                return False

                        else:

                            workload["settings_json"][setting_count]['filters'][filter_option_key].append(filter_part)

            elif filter_option_key != "Node_Type":

                workload['settings_json'][setting_count]['filters'][filter_option_key] = list()

    return True


def check_errors(response):

    error_string = ''

    if "errors" not in response:
        return 'Server Error. Please check this workload set'

    if response["errors"]:
        for error in response["errors"]:
            if error["result_name"] == "Lowest_Cost":
                try:
                    error_string = str(error['message'])
                except ValueError:
                    error_string = "Unknown error. Could not be converted into string"
                break

    return error_string


def binary_search(key, start_range, end_range, workload):

    while end_range - start_range != 1:

        max_value = int((start_range + end_range) / 2)

        if workload['wl_list'][0]['wl_type'] == "VDI_INFRA":
            workload['wl_list'][0]['vm_details']['vm1']['num_vms'] = max_value
        else:
            workload['wl_list'][0][key] = max_value

        post_response = requests.post(host_url, data=json.dumps(workload), headers=header).json()
        response = requests.get(host_url, headers=header).json()

        if check_errors(response):
            end_range = max_value
        else:
            start_range = max_value

    if workload['wl_list'][0]['wl_type'] == "VDI_INFRA":
        workload['wl_list'][0]['vm_details']['vm1']['num_vms'] = end_range
    else:
        workload['wl_list'][0][key] = end_range

    post_response = requests.post(host_url, data=json.dumps(workload), headers=header).json()
    response = requests.get(host_url, headers=header).json()

    if check_errors(response):

        if workload['wl_list'][0]['wl_type'] == "VDI_INFRA":
            workload['wl_list'][0]['vm_details']['vm1']['num_vms'] = start_range
        else:
            workload['wl_list'][0][key] = start_range

        post_response = requests.post(host_url, data=json.dumps(workload), headers=header).json()
        response = requests.get(host_url, headers=header).json()

        return True, response, start_range

    else:

        return True, response, end_range


def get_response(workload, start_range, end_range, configured):

    if configured:
        post_response = requests.post(host_url, data=json.dumps(workload), headers=header).json()
        response = requests.get(host_url, headers=header).json()
        error_string = check_errors(response)
        if error_string:
            return False, error_string, None
        return True, response, None
    else:
        if workload["wl_list"][0]["wl_type"] == "VDI":
            key = "num_desktops"
        elif workload["wl_list"][0]["wl_type"] == "RDSH":
            key = "total_users"
        elif workload["wl_list"][0]["wl_type"] in ["VSI", "ROBO"]:
            key = "num_vms"
        elif workload["wl_list"][0]["wl_type"] in ["DB", "ORACLE"]:
            key = "num_db_instances"

        workload['wl_list'][0][key] = start_range

        post_response = requests.post(host_url, data=json.dumps(workload), headers=header).json()
        response = requests.get(host_url, headers=header).json()

        error_string = check_errors(response)

        if error_string:
            return False, error_string, None
        return binary_search(key, start_range, end_range, workload)


def util_dict_to_string(util_list):

    util_string = ''

    for util_dict, _ in util_list:

        if len(util_list) > 1:
            util_string += '['

        for tag in ['CPU', 'RAM', 'Storage Capacity', 'GPU Users', 'Cache', 'Storage IOPS']:
            util_values = util_dict[tag]
            util_string += '{' + tag + ':' + '[' + str(util_values['wl_util']) + ', ' + str(util_values['ft_util']) + \
                           ']}'

        if len(util_list) > 1:
            util_string += ']'

    return util_string


def write_data_to_excel(sheet, row, cell_data, results):

    if cell_data["Result"]['data'] == 'Fail':
        fill = yellowfill
    else:
        fill = whitefill

    for key in cell_data:

        if key in results or key in ['Result', 'Notes']:
            sheet.cell(row=row, column=cell_data[key]["column"]).value = str(cell_data[key]["data"])

        if ' -> ' in str(cell_data[key]['data']):
            sheet.cell(row=row, column=cell_data[key]["column"]).fill = redfill
        else:
            sheet.cell(row=row, column=cell_data[key]["column"]).fill = fill


def get_node_details(node_info, template_dict):

    for node in node_info:

        num_nodes = node["num_nodes"]

        node_details = node["model_details"]

        node_name = node_details["name"]

        node_description = ""

        for key in ["CPU", "RAM", "HDD", "SSD"]:

            if key in node_details["node_description"]:
                node_description += node_details["node_description"][key]
            else:
                continue
            if key != 'SSD':
                node_description += " | "

        if node['hercules_conf']:
            node_description += " | HyperFlex Acceleration Engine"

        if node_details["subtype"] != "compute":

            if template_dict["HX Desc"] == '--':
                template_dict["HX Desc"] = node_description
            elif template_dict["HX Desc"].endswith(']'):
                template_dict["HX Desc"] += '[' + node_description + ']'
            else:
                template_dict["HX Desc"] = '[' + template_dict["HX Desc"] + '][' + node_description + ']'

            if template_dict["HX Count"]:
                count_string = template_dict["HX Count"] + ', '
            else:
                count_string = ''

            if template_dict["HX Node"] and template_dict["HX Node"] != '--':
                node_string = template_dict["HX Node"] + ', '
            else:
                node_string = ''

            template_dict["HX Count"] = count_string + str(num_nodes)

            template_dict["HX Node"] = node_string + str(node_name)

        else:

            if template_dict["CO Desc"] == '--':
                template_dict["CO Desc"] = node_description
            elif template_dict["CO Desc"].endswith(']'):
                template_dict["CO Desc"] += '[' + node_description + ']'
            else:
                template_dict["CO Desc"] = '[' + template_dict["CO Desc"] + '][' + node_description + ']'

            if template_dict["CO Count"]:
                count_string = template_dict["CO Count"] + ', '
            else:
                count_string = ''

            if template_dict["CO Node"] and template_dict["CO Node"] != '--':
                node_string = template_dict["CO Node"] + ', '
            else:
                node_string = ''

            template_dict["CO Count"] = count_string + str(num_nodes)

            template_dict["CO Node"] = node_string + str(node_name)


def get_brief_util(util_info):

    utilization = dict()

    for util_dict in util_info:
        if not util_dict['status']:
            utilization[util_dict['tag_name']] = {'wl_util': 0,
                                                  'ft_util': 0}
            continue
        utilization[util_dict['tag_name']] = {'wl_util': util_dict['wl_util'],
                                              'ft_util': util_dict['ft_util']}

    return utilization


def check_utilization_error(cell_map, template_dict):

    def get_check_list(wl_list):

        check_list = ['CPU', 'RAM', 'Storage Capacity', 'GPU Users', 'Cache', 'Storage IOPS']

        wl_types = {wl['wl_type'] for wl in wl_list}

        if 'VDI' not in wl_types or not any(True if wl.get('gpu_users', 0) else False for wl in wl_list):
            check_list.remove('GPU Users')

        if 'RAW' in wl_types or 'EXCHANGE' in wl_types:
            check_list.remove('Cache')
            check_list.remove('Storage IOPS')

        if ('VDI_INFRA' in wl_types or 'RDSH' in wl_types) and 'VDI' not in wl_types:
            check_list.remove('Cache')
            check_list.remove('Storage IOPS')

        if 'VEEAM' in wl_types:
            check_list.remove('CPU')
            check_list.remove('RAM')
            check_list.remove('Storage IOPS')
            check_list.remove('Cache')

        if 'SPLUNK' in wl_types:
            check_list.remove('Cache')

        if 'AIML' in wl_types:
            check_list.remove('Storage Capacity')
            check_list.remove('Storage IOPS')
            check_list.remove('Cache')

        return check_list

    util_string = util_dict_to_string(template_dict['Utilization'])

    if any([tag_dict['wl_util'] > 100 or tag_dict['ft_util'] > 100 or (tag_name in get_check_list(wl_list) and
                                                                       (tag_dict['wl_util'] <= 0 or
                                                                        tag_dict['ft_util'] <= 0))
            for util_dict, wl_list in template_dict['Utilization'] for tag_name, tag_dict in util_dict.items()]):

        cell_map["Notes"]["data"] = 'Utilization error has occurred'

    if cell_map['Utilization']['data'] and util_string != cell_map['Utilization']['data']:

        util_string = cell_map['Utilization']['data'] + ' -> ' + util_string

        if cell_map["Notes"]["data"]:
            cell_map["Notes"]["data"] += ', '

        cell_map["Notes"]["data"] += 'Utilization regression has occurred'

    cell_map['Utilization']['data'] = util_string


def get_regressions(result_list, cell_map, template_dict):

    for key in result_list:

        if key == 'Utilization':

            check_utilization_error(cell_map, template_dict)

        elif cell_map[key]["data"] and str(cell_map[key]["data"]) != str(template_dict[key]):

            cell_map[key]["data"] = str(cell_map[key]["data"]) + " -> " + str(template_dict[key])

            if cell_map["Notes"]["data"]:
                cell_map["Notes"]["data"] += ', '

            cell_map["Notes"]["data"] += 'Regressions have occurred'

        else:

            cell_map[key]["data"] = str(template_dict[key])


def construct_result(status, response, max_value, cell_map, row_num, sheet, wl_list):

    sheet_name = sheet.title

    result_set = RESULT_SET.intersection(cell_map.keys())

    error_type = 'regressions'

    cell_map["Notes"]['data'] = ""

    if not status:
        for result_key in result_set:

            cell_map["Notes"]["data"] = response

            if cell_map[result_key]["data"] and cell_map[result_key]["data"] != '--':

                cell_map[result_key]["data"] = str(cell_map[result_key]["data"]) + " -> --"

            else:

                cell_map[result_key]["data"] = "--"

                error_type = 'expected errors'

    else:

        template_dict = {"Max #": max_value if max_value else "--",
                         "HX Count": 0,
                         "HX Node": "--",
                         "HX Desc": "--",
                         "CO Count": 0,
                         "CO Node": "--",
                         "CO Desc": "--",
                         "Accessory Count": 0,
                         "Accessories": "--",
                         "Utilization": "--"}

        for result_dict in response["workload_result"]:

            if result_dict["result_name"] == "Lowest_Cost":

                template_dict['Utilization'] = list()

                for cluster_info in result_dict["clusters"]:

                    for cluster_node in cluster_info:

                        node_info = cluster_node["node_info"]

                        get_node_details(node_info, template_dict)

                        template_dict['Utilization'].append([get_brief_util(cluster_node["Utilization"]),
                                                            cluster_node['wl_list']])

                        for accessory in cluster_node["accessories"]:

                            if template_dict["Accessory Count"]:
                                count_string = template_dict["Accessory Count"] + ', '
                            else:
                                count_string = ''

                            if template_dict["Accessories"] and template_dict["Accessories"] != "--":
                                accessory_string = template_dict["Accessories"] + ', '
                            else:
                                accessory_string = ''

                            template_dict["Accessory Count"] = count_string + str(accessory["count"])

                            template_dict["Accessories"] = accessory_string + str(accessory["part_name"])

                break

        get_regressions(result_set, cell_map, template_dict)

    if cell_map["Notes"]['data'] or any(' -> ' in str(cell_map[key]['data']) for key in cell_map):

        error_count[sheet_name][error_type] += 1

        cell_map["Result"]['data'] = 'Fail'

    else:

        cell_map["Result"]['data'] = 'Pass'

    write_data_to_excel(sheet, row_num, cell_map, result_set)


def configure_default_profile(workload, wl_type, profile, workload_index):

    if wl_type == 'INFRA':

        workload['wl_list'][workload_index]['vcpus_per_core'] = 1
        workload['wl_list'][workload_index]['ram_opratio'] = 1

        for _, vm in workload['wl_list'][workload_index]['vm_details'].items():
            vm['vcpus_per_vm'] = 4
            vm['ram_per_vm'] = 8
            vm['disk_per_vm'] = 32

        return

    elif wl_type == 'VDI':

        if profile == "Task Worker":
            profile_dict = {"vcpus_per_desktop": 1,
                            "clock_per_desktop":
                                325 if workload['wl_list'][workload_index]["os_type"] == "win_7" else 370,
                            "ram_per_desktop": 1,
                            "avg_iops_per_desktop": 6,
                            "gold_image_size": 20 if workload['wl_list'][workload_index]["os_type"] == "win_7" else 40,
                            "gold_image_size_unit": 'GB',
                            "snapshots":
                                0 if workload['wl_list'][workload_index]["provisioning_type"] == "Pooled Desktops" else 5,
                            "working_set": 10}
        elif profile == "Power User":
            profile_dict = {"vcpus_per_desktop": 2,
                            "clock_per_desktop":
                                400 if workload['wl_list'][workload_index]["os_type"] == "win_7" else 500,
                            "ram_per_desktop": 2 if workload['wl_list'][workload_index]["os_type"] == "win_7" else 4,
                            "avg_iops_per_desktop": 10,
                            "gold_image_size": 20 if workload['wl_list'][workload_index]["os_type"] == "win_7" else 40,
                            "gold_image_size_unit": 'GB',
                            "snapshots":
                                0 if workload['wl_list'][workload_index]["provisioning_type"] == "Pooled Desktops" else 5,
                            "working_set": 10}
        elif profile == "Knowledge Worker":
            profile_dict = {"vcpus_per_desktop": 2,
                            "clock_per_desktop":
                                400 if workload['wl_list'][workload_index]["os_type"] == "win_7" else 500,
                            "ram_per_desktop": 2 if workload['wl_list'][workload_index]["os_type"] == "win_7" else 4,
                            "avg_iops_per_desktop": 8,
                            "gold_image_size": 20 if workload['wl_list'][workload_index]["os_type"] == "win_7" else 40,
                            "gold_image_size_unit": 'GB',
                            "snapshots":
                                0 if workload['wl_list'][workload_index]["provisioning_type"] == "Pooled Desktops" else 5,
                            "working_set": 10}

    elif wl_type == 'RDSH':

        if profile == "Task Worker":
            profile_dict = {"vcpus_per_vm": 2,
                            "sessions_per_vm":
                                25 if workload['wl_list'][workload_index]["broker_type"] == "citrix" else 30,
                            "clock_per_session": 325,
                            "max_vcpus_per_core": 1,
                            "ram_per_vm": 1,
                            "os_per_vm": 20,
                            "os_per_vm_unit": 'GB'}
        elif profile == "Power User":
            profile_dict = {"vcpus_per_vm": 2,
                            "sessions_per_vm":
                                25 if workload['wl_list'][workload_index]["broker_type"] == "citrix" else 30,
                            "clock_per_session": 375,
                            "max_vcpus_per_core": 1,
                            "ram_per_vm": 1,
                            "os_per_vm": 20,
                            "os_per_vm_unit": 'GB'}
        elif profile == "Knowledge Worker":
            profile_dict = {"vcpus_per_vm": 2,
                            "sessions_per_vm":
                                25 if workload['wl_list'][workload_index]["broker_type"] == "citrix" else 30,
                            "clock_per_session": 400,
                            "max_vcpus_per_core": 1,
                            "ram_per_vm": 1,
                            "os_per_vm": 20,
                            "os_per_vm_unit": 'GB'}

    elif wl_type == "VSI/ROBO":

        if profile == "Small":
            profile_dict = {"vcpus_per_vm": 2,
                            "vcpus_per_core": 4,
                            "ram_per_vm": 8,
                            "avg_iops_per_vm": 50,
                            "disk_per_vm": 50,
                            "disk_per_vm_unit": 'GB',
                            "base_image_size": 20,
                            "base_image_size_unit": 'GB',
                            "snapshots": 5,
                            "working_set": 10}
        elif profile == "Medium":
            profile_dict = {"vcpus_per_vm": 4,
                            "vcpus_per_core": 4,
                            "ram_per_vm": 16,
                            "avg_iops_per_vm": 100,
                            "disk_per_vm": 200,
                            "disk_per_vm_unit": 'GB',
                            "base_image_size": 20,
                            "base_image_size_unit": 'GB',
                            "snapshots": 5,
                            "working_set": 10}
        elif profile == "Large":
            profile_dict = {"vcpus_per_vm": 8,
                            "vcpus_per_core": 4,
                            "ram_per_vm": 32,
                            "avg_iops_per_vm": 200,
                            "disk_per_vm": 750,
                            "disk_per_vm_unit": 'GB',
                            "base_image_size": 20,
                            "base_image_size_unit": 'GB',
                            "snapshots": 5,
                            "working_set": 10}

    elif wl_type == "SQL":

        iops_type = "avg_iops_per_db" if workload["wl_list"][workload_index]["db_type"] == "OLTP" else "avg_mbps_per_db"

        iops_dict = {"Small": {"OLTP": 1000,
                               "OLAP": 100},
                     "Medium": {"OLTP": 3000,
                                "OLAP": 200},
                     "Large": {"OLTP": 10000,
                               "OLAP": 800}}

        if profile == "Small":
            profile_dict = {"vcpus_per_db": 2,
                            "vcpus_per_core": 2,
                            "ram_per_db": 8,
                            "db_size": 400,
                            "db_size_unit": 'GB',
                            iops_type: iops_dict[profile][workload["wl_list"][workload_index]["db_type"]],
                            "metadata_size": 45}
        elif profile == "Medium":
            profile_dict = {"vcpus_per_db": 4,
                            "vcpus_per_core": 2,
                            "ram_per_db": 16,
                            "db_size": 1000,
                            "db_size_unit": 'GB',
                            iops_type: iops_dict[profile][workload["wl_list"][workload_index]["db_type"]],
                            "metadata_size": 40}
        elif profile == "Large":
            profile_dict = {"vcpus_per_db": 8,
                            "vcpus_per_core": 2,
                            "ram_per_db": 32,
                            "db_size": 4000,
                            "db_size_unit": 'GB',
                            iops_type: iops_dict[profile][workload["wl_list"][workload_index]["db_type"]],
                            "metadata_size": 30}

    elif wl_type == "ORACLE":

        iops_type = "avg_iops_per_db" if workload["wl_list"][workload_index]["db_type"] == "OLTP" else "avg_mbps_per_db"

        iops_dict = {"Small": {"OLTP": 6000,
                               "OLAP": 200},
                     "Medium": {"OLTP": 10000,
                                "OLAP": 400},
                     "Large": {"OLTP": 30000,
                               "OLAP": 1000}}

        if profile == "Small":
            profile_dict = {"vcpus_per_db": 4,
                            "vcpus_per_core": 2,
                            "ram_per_db": 16,
                            "db_size": 400,
                            "db_size_unit": 'GB',
                            iops_type: iops_dict[profile][workload["wl_list"][workload_index]["db_type"]],
                            "metadata_size": 45}
        elif profile == "Medium":
            profile_dict = {"vcpus_per_db": 8,
                            "vcpus_per_core": 2,
                            "ram_per_db": 64,
                            "db_size": 1000,
                            "db_size_unit": 'GB',
                            iops_type: iops_dict[profile][workload["wl_list"][workload_index]["db_type"]],
                            "metadata_size": 40}
        elif profile == "Large":
            profile_dict = {"vcpus_per_db": 16,
                            "vcpus_per_core": 2,
                            "ram_per_db": 96,
                            "db_size": 4000,
                            "db_size_unit": 'GB',
                            iops_type: iops_dict[profile][workload["wl_list"][workload_index]["db_type"]],
                            "metadata_size": 30}

    else:
        return

    for key, value in profile_dict.items():
        workload['wl_list'][workload_index][key] = value


def configure_custom_profile(workload, wl_type, data, workload_index):

    if wl_type == "DB/ORACLE":

        db_size_unit = data['DB size Unit']['data'] if data['DB size Unit']['data'] else 'GB'

        iops_type = "avg_iops_per_db" if workload["wl_list"][workload_index]["db_type"] == "OLTP" else "avg_mbps_per_db"

        profile_dict = {"vcpus_per_db": int(data['vCPU']['data']),
                        "vcpus_per_core": int(data['vCPU/core']['data']),
                        "ram_per_db": float(data['RAM']['data']),
                        "db_size": float(data['DB size']['data']),
                        "db_size_unit": db_size_unit,
                        iops_type: int(data['IOPS']['data']),
                        "metadata_size": int(data['DB overhead']['data'])}

    elif wl_type == "VDI":

        profile_dict = {"vcpus_per_desktop": int(data['vCPU']['data']),
                        "clock_per_desktop": int(data['Clock']['data']),
                        "ram_per_desktop": float(data['RAM']['data']),
                        "avg_iops_per_desktop": int(data['OS IOPS']['data']),
                        "gold_image_size": float(data['OS size']['data']),
                        "gold_image_size_unit":
                            data['OS size unit']['data']
                            if "OS size unit" in data and data['OS size unit']['data'] else 'GB',
                        "snapshots": int(data['Snapshots']['data']),
                        "working_set": int(data['Working set']['data'])}

    elif wl_type == "RDSH":

        profile_dict = {"vcpus_per_vm": int(data['vCPU']['data']),
                        "sessions_per_vm": int(data['Sessions']['data']),
                        "clock_per_session": int(data['Clock']['data']),
                        "max_vcpus_per_core": float(data['Max OP ratio']['data']),
                        "ram_per_vm": float(data['RAM']['data']),
                        "os_per_vm": float(data['OS size']['data']),
                        "os_per_vm_unit":
                            data['OS size unit']['data']
                            if "OS size unit" in data and data['OS size unit']['data'] else 'GB'}

    elif wl_type == "VSI/ROBO":

        disk_per_vm_unit = data['Disk Unit']['data'] if 'Disk Unit' in data and data['Disk Unit']['data'] else 'GB'

        profile_dict = {"vcpus_per_vm": int(data['vCPU']['data']),
                        "vcpus_per_core": int(data['vCPU/core']['data']),
                        "ram_per_vm": float(data['RAM']['data']),
                        "avg_iops_per_vm": int(data['IOPS']['data']),
                        "disk_per_vm": float(data['Disk']['data']),
                        "disk_per_vm_unit": disk_per_vm_unit,
                        "base_image_size": float(data['Base image size']['data']),
                        "base_image_size_unit":
                            data['Base image size unit']['data']
                            if 'Base image size unit' in data and data['Base image size unit']['data'] else 'GB',
                        "snapshots": int(data['Snapshots']['data']),
                        "working_set": int(data['Working set']['data'])}

    for key, value in profile_dict.items():
        workload['wl_list'][workload_index][key] = value


def test_vdi(sheet, workload, cell_map, workload_set_map):

    template_wl = deepcopy(workload['wl_list'][0])

    for element in workload_set_map:

        workload['wl_list'] = list()

        for row_num in range(element[0], element[1] + 1):

            workload['wl_list'].append(deepcopy(template_wl))

            workload_index = row_num - element[0]

            for key, value_dict in cell_map.items():
                if key in RESULT_SET and row_num != element[0]:
                    continue
                value_dict["data"] = sheet.cell(row=row_num, column=value_dict["column"]).value

            workload['wl_list'][workload_index]['os_type'] = CONVERSION_DICT[cell_map['OS']['data']]

            workload['wl_list'][workload_index]['profile_type'] = CONVERSION_DICT[cell_map['User']['data']]

            workload['wl_list'][workload_index]['provisioning_type'] = CONVERSION_DICT[cell_map['Provisioning']["data"]]

            workload['wl_list'][workload_index]['replication_factor'] = int(cell_map['RF']['data'])

            workload['wl_list'][workload_index]['fault_tolerance'] = int(cell_map['FT']["data"])

            workload['wl_list'][workload_index]['wl_name'] = sheet.title + ' - ' + str(row_num)

            workload['wl_list'][workload_index]['user_iops'] = \
                int(cell_map['User IOPS']["data"]) if cell_map['User IOPS']["data"] else 0

            workload['wl_list'][workload_index]['disk_per_desktop'] = \
                float(cell_map['User data size']["data"]) if cell_map['User data size']["data"] else 0

            workload['wl_list'][workload_index]['disk_per_desktop_unit'] = \
                cell_map['User data size unit']["data"] if cell_map['User data size unit']["data"] else 'GB'

            workload['wl_list'][workload_index]['vdi_directory'] = \
                True if cell_map['Home Directory']['data'] == 'yes' else False

            if cell_map["GPU"]["data"] == "Enabled":

                workload['wl_list'][workload_index]["gpu_users"] = 1

                if "video_RAM" in cell_map and cell_map["video_RAM"]["data"]:
                    workload['wl_list'][workload_index]["video_RAM"] = int(cell_map["video_RAM"]["data"])
                else:
                    workload['wl_list'][workload_index]["video_RAM"] = 1024
            else:
                workload['wl_list'][workload_index]["gpu_users"] = 0

            configured = False

            if cell_map["#Desktops"]["data"]:

                workload['wl_list'][workload_index]['num_desktops'] = int(cell_map["#Desktops"]["data"])

                configured = True

                end_range = None

                start_range = None

            else:

                start_range = 1

                end_range = MAX_LIMIT["VDI"]

            if not workload_index:
                status = edit_settings(workload, cell_map)
            if not status:
                construct_result(status, 'Wrong inputs have been given.', None, cell_map, element[0], sheet,
                                 workload['wl_list'])
                break

            if workload['wl_list'][workload_index]['profile_type'] != "Custom User":
                configure_default_profile(workload, "VDI", workload['wl_list'][workload_index]['profile_type'],
                                          workload_index)
            else:
                configure_custom_profile(workload, "VDI", cell_map, workload_index)

        try:
            status, response, max_value = get_response(workload, start_range, end_range, configured)
        except ValueError:
            status, response, max_value = False, 'Server Error. Please check this workload set', None

        construct_result(status, response, max_value, cell_map, element[0], sheet, workload['wl_list'])


def test_rdsh(sheet, workload, cell_map, workload_set_map):

    template_wl = deepcopy(workload['wl_list'][0])

    for element in workload_set_map:

        workload['wl_list'] = list()

        for row_num in range(element[0], element[1] + 1):

            workload['wl_list'].append(deepcopy(template_wl))

            workload_index = row_num - element[0]

            for key, value_dict in cell_map.items():
                if key in RESULT_SET and row_num != element[0]:
                    continue
                value_dict["data"] = sheet.cell(row=row_num, column=value_dict["column"]).value

            workload['wl_list'][workload_index]['profile_type'] = CONVERSION_DICT[cell_map['User']['data']]

            workload['wl_list'][workload_index]['broker_type'] = CONVERSION_DICT[cell_map['Broker']['data']]

            workload['wl_list'][workload_index]['replication_factor'] = int(cell_map['RF']['data'])

            workload['wl_list'][workload_index]['fault_tolerance'] = int(cell_map['FT']["data"])

            workload['wl_list'][workload_index]['wl_name'] = sheet.title + ' - ' + str(row_num)

            workload['wl_list'][workload_index]['user_iops'] = \
                int(cell_map['User IOPS']["data"]) if cell_map['User IOPS']["data"] else 0

            workload['wl_list'][workload_index]['hdd_per_user'] = \
                float(cell_map['User data size']["data"]) if cell_map['User data size']["data"] else 0

            workload['wl_list'][workload_index]['hdd_per_user_unit'] = \
                cell_map['User data size unit']["data"] if cell_map['User data size unit']["data"] else 'GB'

            workload['wl_list'][workload_index]['rdsh_directory'] = \
                True if cell_map['Home Directory']['data'] == 'yes' else False

            if cell_map["GPU"]["data"] == "Enabled":

                workload['wl_list'][workload_index]["gpu_users"] = 1

                if "video_RAM" in cell_map and cell_map["video_RAM"]["data"]:
                    workload['wl_list'][workload_index]["video_RAM"] = int(cell_map["video_RAM"]["data"])
                else:
                    workload['wl_list'][workload_index]["video_RAM"] = 8192
            else:
                workload['wl_list'][workload_index]["gpu_users"] = 0

            configured = False

            if cell_map["#Users"]["data"]:

                workload['wl_list'][workload_index]['total_users'] = int(cell_map["#Users"]["data"])

                configured = True

                end_range = None

                start_range = None

            else:

                start_range = 1

                end_range = MAX_LIMIT["RDSH"]

            if not workload_index:
                status = edit_settings(workload, cell_map)
            if not status:
                construct_result(status, 'Wrong inputs have been given.', None, cell_map, element[0], sheet,
                                 workload['wl_list'])
                break

            if workload['wl_list'][workload_index]['profile_type'] != "Custom User":
                configure_default_profile(workload, "RDSH", workload['wl_list'][workload_index]['profile_type'],
                                          workload_index)
            else:
                configure_custom_profile(workload, "RDSH", cell_map, workload_index)

        try:
            status, response, max_value = get_response(workload, start_range, end_range, configured)
        except ValueError:
            status, response, max_value = False, 'Server Error. Please check this workload set', None

        construct_result(status, response, max_value, cell_map, element[0], sheet, workload['wl_list'])

def test_aiml(sheet, workload, cell_map, workload_set_map):

    template_wl = deepcopy(workload['wl_list'][0])

    for element in workload_set_map:

        workload['wl_list'] = list()

        for row_num in range(element[0], element[1] + 1):

            workload['wl_list'].append(deepcopy(template_wl))
            workload_index = row_num - element[0]

            for key, value_dict in cell_map.items():
                if key in RESULT_SET and row_num != element[0]:
                    continue
                value_dict["data"] = sheet.cell(row=row_num, column=value_dict["column"]).value

            if not workload_index:
                status = edit_settings(workload, cell_map)
            if not status:
                construct_result(status, 'Wrong inputs have been given.', None, cell_map, element[0], sheet,
                                 workload['wl_list'])
                break

            workload['wl_list'][workload_index]['wl_name'] = sheet.title + ' - ' + str(row_num)

            workload['wl_list'][workload_index]['replication_factor'] = int(cell_map['RF']['data'])

            workload['wl_list'][workload_index]['fault_tolerance'] = int(cell_map['FT']["data"])

            workload['wl_list'][workload_index]['input_type'] = cell_map['Input_Source']['data']

            workload['wl_list'][workload_index]['expected_util'] = cell_map['Expected_Utilization']['data']

            workload['wl_list'][workload_index]['num_data_scientists'] = int(cell_map['Number_DataScientist']['data'])

            workload['wl_list'][workload_index]['vcpus_per_ds'] = int(cell_map['Cores_Per_DS']['data'])

            workload['wl_list'][workload_index]['vcpus_per_core'] = int(cell_map['vCPU/core']['data'])

            workload['wl_list'][workload_index]['ram_per_ds'] = int(cell_map['RAM_Per_DS']['data'])

            workload['wl_list'][workload_index]['gpu_per_ds'] = int(cell_map['GPU_Per_DS']['data'])

            workload['wl_list'][workload_index]['enablestorage'] = \
                True if cell_map['Storage_HXcluster']['data'] == 'yes' else False

            workload['wl_list'][workload_index]['disk_per_ds'] = int(cell_map['Disk_Per_DS']['data'])

            workload['wl_list'][workload_index]['disk_per_ds_unit'] = cell_map['Disk_Unit']['data']


        try:
            status, response, max_value = get_response(workload, None, None, True)
        except ValueError:
            status, response, max_value = False, 'Server Error. Please check this workload set', None

        construct_result(status, response, max_value, cell_map, element[0], sheet, workload['wl_list'])


def test_vsi_robo(sheet, workload, cell_map, workload_set_map):

    template_wl = deepcopy(workload['wl_list'][0])

    for element in workload_set_map:

        workload['wl_list'] = list()

        for row_num in range(element[0], element[1] + 1):

            workload['wl_list'].append(deepcopy(template_wl))

            workload_index = row_num - element[0]

            for key, value_dict in cell_map.items():
                if key in RESULT_SET and row_num != element[0]:
                    continue
                value_dict["data"] = sheet.cell(row=row_num, column=value_dict["column"]).value

            workload['wl_list'][workload_index]['profile_type'] = cell_map['Type']['data']

            workload['wl_list'][workload_index]['replication_factor'] = int(cell_map['RF']['data'])

            workload['wl_list'][workload_index]['fault_tolerance'] = int(cell_map['FT']["data"])

            workload['wl_list'][workload_index]['wl_name'] = sheet.title + ' - ' + str(row_num)

            configured = False

            if cell_map["#VMs"]["data"]:
                workload['wl_list'][workload_index]['num_vms'] = int(cell_map["#VMs"]["data"])
                configured = True
                end_range = None
                start_range = None
            else:
                start_range = 1
                end_range = MAX_LIMIT[workload['wl_list'][0]['wl_type']]

            if not workload_index:
                status = edit_settings(workload, cell_map)
            if not status:
                construct_result(status, 'Wrong inputs have been given.', None, cell_map, element[0], sheet,
                                 workload['wl_list'])
                break

            if workload['wl_list'][workload_index]['profile_type'] != "Custom":
                configure_default_profile(workload, "VSI/ROBO",
                                          workload['wl_list'][workload_index]['profile_type'], workload_index)
            else:
                configure_custom_profile(workload, "VSI/ROBO", cell_map, workload_index)

        try:
            status, response, max_value = get_response(workload, start_range, end_range, configured)
        except ValueError:
            status, response, max_value = False, 'Server Error. Please check this workload set', None

        construct_result(status, response, max_value, cell_map, element[0], sheet, workload['wl_list'])


def test_db(sheet, workload, cell_map, workload_set_map):

    template_wl = deepcopy(workload['wl_list'][0])

    for element in workload_set_map:

        workload['wl_list'] = list()

        for row_num in range(element[0], element[1] + 1):

            workload['wl_list'].append(deepcopy(template_wl))

            workload_index = row_num - element[0]

            for key, value_dict in cell_map.items():
                if key in RESULT_SET and row_num != element[0]:
                    continue
                value_dict["data"] = sheet.cell(row=row_num, column=value_dict["column"]).value

            workload['wl_list'][workload_index]['profile_type'] = cell_map['Type']['data']

            workload['wl_list'][workload_index]['db_type'] = cell_map['DB type']['data']

            workload['wl_list'][workload_index]['wl_name'] = sheet.title + ' - ' + str(row_num)

            workload['wl_list'][workload_index]['replication_factor'] = int(cell_map['RF']['data'])

            workload['wl_list'][workload_index]['fault_tolerance'] = int(cell_map['FT']["data"])

            if workload['wl_list'][workload_index]['wl_type'] == 'ORACLE':
                workload['wl_list'][workload_index]['internal_type'] = \
                    "O" + workload['wl_list'][workload_index]['db_type']
                DB = 'ORACLE'
            else:
                workload['wl_list'][workload_index]['internal_type'] = workload['wl_list'][workload_index]['db_type']
                DB = 'SQL'

            configured = False

            if cell_map["#DBs"]["data"]:
                workload['wl_list'][workload_index]['num_db_instances'] = int(cell_map["#DBs"]["data"])
                configured = True
                end_range = None
                start_range = None
            else:
                start_range = 1
                end_range = MAX_LIMIT[workload['wl_list'][workload_index]['wl_type']]

            if not workload_index:
                status = edit_settings(workload, cell_map)
            if not status:
                construct_result(status, 'Wrong inputs have been given.', None, cell_map, element[0], sheet,
                                 workload['wl_list'])
                break

            if workload['wl_list'][workload_index]['profile_type'] != "Custom":
                configure_default_profile(workload, DB, workload['wl_list'][workload_index]['profile_type'],
                                          workload_index)
            else:
                configure_custom_profile(workload, 'DB/ORACLE', cell_map, workload_index)

        try:
            status, response, max_value = get_response(workload, start_range, end_range, configured)
        except ValueError:
            status, response, max_value = False, 'Server Error. Please check this workload set', None

        construct_result(status, response, max_value, cell_map, element[0], sheet, workload['wl_list'])


def test_infra(sheet, workload, cell_map, workload_set_map):

    template_wl = deepcopy(workload['wl_list'][0])

    for element in workload_set_map:

        workload['wl_list'] = list()

        for row_num in range(element[0], element[1] + 1):

            workload['wl_list'].append(deepcopy(template_wl))
            workload_index = row_num - element[0]

            for key, value_dict in cell_map.items():
                if key in RESULT_SET and row_num != element[0]:
                    continue
                value_dict["data"] = sheet.cell(row=row_num, column=value_dict["column"]).value

            if not workload_index:
                status = edit_settings(workload, cell_map)
            if not status:
                construct_result(status, 'Wrong inputs have been given.', None, cell_map, element[0], sheet,
                                 workload['wl_list'])
                break

            workload['wl_list'][workload_index]['wl_name'] = sheet.title + ' - ' + str(row_num)

            workload['wl_list'][workload_index]['replication_factor'] = int(cell_map['RF']['data'])

            workload['wl_list'][workload_index]['fault_tolerance'] = int(cell_map['FT']["data"])

            workload['wl_list'][workload_index]['ram_opratio'] = \
                int(cell_map["Ram OP"]["data"]) if cell_map["Ram OP"]["data"] else 1

            workload['wl_list'][workload_index]['vcpus_per_core'] = \
                int(cell_map["CPU OP"]["data"]) if cell_map["CPU OP"]["data"] else 1

            workload['wl_list'][workload_index]['vm_details']['vm1']['vcpus_per_vm'] = int(cell_map['vCPU-1']['data'])

            workload['wl_list'][workload_index]['vm_details']['vm1']['ram_per_vm'] = int(cell_map['RAM-1']['data'])

            workload['wl_list'][workload_index]['vm_details']['vm1']['disk_per_vm'] = int(cell_map['Disk-1']['data'])

            workload['wl_list'][workload_index]['vm_details']['vm2']['vcpus_per_vm'] = \
                int(cell_map['vCPU-2']['data']) if cell_map['vCPU-2']['data'] else 1

            workload['wl_list'][workload_index]['vm_details']['vm2']['ram_per_vm'] = \
                int(cell_map['RAM-2']['data']) if cell_map['RAM-2']['data'] else 1

            workload['wl_list'][workload_index]['vm_details']['vm2']['disk_per_vm'] = \
                int(cell_map['Disk-2']['data']) if cell_map['Disk-2']['data'] else 1

            if not workload_index and cell_map['test-type']['data'] == 'normal':

                workload['wl_list'][workload_index]['vm_details']['vm1']['num_vms'] = int(cell_map['#VMs - 1']['data'])
                workload['wl_list'][workload_index]['vm_details']['vm2']['num_vms'] = int(cell_map['#VMs - 2']['data'])

                try:
                    status, response, max_value = get_response(workload, None, None, True)
                except ValueError:
                    status, response, max_value = False, 'Server Error. Please check this workload set', None

                construct_result(status, response, max_value, cell_map, element[0], sheet, workload['wl_list'])

            elif not workload_index and cell_map['test-type']['data'] == 'maximum':

                workload['wl_list'][workload_index]['vm_details']['vm2']['num_vms'] = 0

                start_range = 1
                end_range = 2000

                try:
                    workload['wl_list'][workload_index]['vm_details']['vm1']['num_vms'] = start_range
                    status, response, _ = get_response(workload, None, None, True)

                    if status:
                        status, response, max_value = binary_search(key, start_range, end_range, workload)
                        constraint_cell = sheet.cell(row=row_num, column=cell_map['#VMs - 1']['column'])

                        if cell_map['#VMs - 1']['data'] and cell_map['#VMs - 1']['data'] != str(max_value):
                            constraint_cell.value = str(cell_map['#VMs - 1']['data']) + ' -> ' + str(max_value)
                        else:
                            constraint_cell.value = str(max_value)

                        construct_result(status, response, None, cell_map, element[0], sheet, workload['wl_list'])

                    else:
                        construct_result(status, response, None, cell_map, element[0], sheet, workload['wl_list'])

                except ValueError:

                    status, response = False, 'Server Error. Please check this workload set'
                    construct_result(status, response, None, cell_map, element[0], sheet, workload['wl_list'])


def test_veeam(sheet, workload, cell_map, workload_set_map):

    template_wl = deepcopy(workload['wl_list'][0])

    for element in workload_set_map:

        workload['wl_list'] = list()

        for row_num in range(element[0], element[1] + 1):

            workload['wl_list'].append(deepcopy(template_wl))
            workload_index = row_num - element[0]

            for key, value_dict in cell_map.items():
                if key in RESULT_SET and row_num != element[0]:
                    continue
                value_dict["data"] = sheet.cell(row=row_num, column=value_dict["column"]).value

            if not workload_index:
                status = edit_settings(workload, cell_map)
            if not status:
                construct_result(status, 'Wrong inputs have been given.', None, cell_map, element[0], sheet,
                                 workload['wl_list'])
                break

            # custom reset of node-options in order to handle VEEAM filter corner case
            for setting in workload['settings_json']:
                setting['filters']['Node_Type'] = list()

            workload['wl_list'][workload_index]['wl_name'] = sheet.title + ' - ' + str(row_num)

            workload['wl_list'][workload_index]['replication_factor'] = int(cell_map['RF']['data'])

            workload['wl_list'][workload_index]['fault_tolerance'] = int(cell_map['FT']["data"])

            workload['wl_list'][workload_index]['hdd_size'] = \
                float(cell_map['HDD_size']['data']) if cell_map['HDD_size']['data'] else 0

            workload['wl_list'][workload_index]['hdd_size_unit'] = cell_map['HDD_size_unit']['data']

        try:

            if cell_map['HDD_size']['data']:
                status, response, max_value = get_response(workload, None, None, True)
                construct_result(status, response, None, cell_map, element[0], sheet, workload['wl_list'])
                continue

            key = 'hdd_size'
            end_range = 3000000
            start_range = 1

            workload['wl_list'][workload_index][key] = start_range
            status, response, _ = get_response(workload, None, None, True)

            if status:
                status, response, max_value = binary_search(key, start_range, end_range, workload)
                constraint_cell = sheet.cell(row=row_num, column=cell_map['HDD_size']['column'])

                if cell_map['HDD_size']['data'] and cell_map['HDD_size']['data'] != str(max_value):
                    constraint_cell.value = str(cell_map['HDD_size']['data']) + ' -> ' + str(max_value)
                else:
                    constraint_cell.value = str(max_value)
            else:
                construct_result(status, response, None, cell_map, element[0], sheet, workload['wl_list'])
                continue

        except ValueError:
            status, response, max_value = False, 'Server Error. Please check this workload set', None

        construct_result(status, response, max_value, cell_map, element[0], sheet, workload['wl_list'])


def test_splunk(sheet, workload, cell_map, workload_set_map):

    template_wl = deepcopy(workload['wl_list'][0])

    for element in workload_set_map:

        workload['wl_list'] = list()

        for row_num in range(element[0], element[1] + 1):

            workload['wl_list'].append(deepcopy(template_wl))
            workload_index = row_num - element[0]

            for key, value_dict in cell_map.items():
                if key in RESULT_SET and row_num != element[0]:
                    continue
                value_dict["data"] = sheet.cell(row=row_num, column=value_dict["column"]).value

            if not workload_index:
                status = edit_settings(workload, cell_map)
            if not status:
                construct_result(status, 'Wrong inputs have been given.', None, cell_map, element[0], sheet,
                                 workload['wl_list'])
                break

            workload['wl_list'][workload_index]['wl_name'] = sheet.title + ' - ' + str(row_num)

            workload['wl_list'][workload_index]['replication_factor'] = int(cell_map['RF']['data'])

            workload['wl_list'][workload_index]['fault_tolerance'] = int(cell_map['FT']["data"])

            workload['wl_list'][workload_index]['daily_data_ingest'] = float(cell_map['Daily_Data_Ingest']['data'])

            workload['wl_list'][workload_index]['daily_data_ingest_unit'] = cell_map['Daily_Data_Ingest_unit']['data']

            workload['wl_list'][workload_index]['max_vol_ind'] = float(cell_map['Max_Vol_Ind']['data'])

            workload['wl_list'][workload_index]['max_vol_ind_unit'] = cell_map['Max_Vol_Ind_unit']['data']

            workload['wl_list'][workload_index]['acc_type'] = CONVERSION_DICT[cell_map['Accumulation_Type']['data']]

            workload['wl_list'][workload_index]['profile_type'] = cell_map['Type']['data']

            storage_dict = workload['wl_list'][workload_index]['storage_acc']

            for stor_key in ['hot', 'warm', 'cold', 'frozen']:

                if cell_map[stor_key]['data']:
                    storage_dict[stor_key] = cell_map[stor_key]['data']

            for vm_name, vm_dict in workload['wl_list'][workload_index]['vm_details'].items():

                vm_dict['vcpus_per_vm'] = cell_map['%s-CPU' % vm_name]['data']

                vm_dict['ram_per_vm'] = cell_map['%s-RAM' % vm_name]['data']

                vm_dict['avg_iops_per_vm'] = cell_map['%s-IOPS' % vm_name]['data']

                vm_dict['num_vms'] = cell_map['%s-count' % vm_name]['data']

        try:
            status, response, max_value = get_response(workload, None, None, True)
        except ValueError:
            status, response, max_value = False, 'Server Error. Please check this workload set', None

        construct_result(status, response, max_value, cell_map, element[0], sheet, workload['wl_list'])


def test_raw(sheet, workload, cell_map, workload_set_map):

    template_wl = deepcopy(workload['wl_list'][0])

    for element in workload_set_map:

        workload['wl_list'] = list()

        for row_num in range(element[0], element[1] + 1):

            workload['wl_list'].append(deepcopy(template_wl))
            workload_index = row_num - element[0]

            for key, value_dict in cell_map.items():
                if key in RESULT_SET and row_num != element[0]:
                    continue
                value_dict["data"] = sheet.cell(row=row_num, column=value_dict["column"]).value

            workload['wl_list'][workload_index]['wl_name'] = sheet.title + ' - ' + str(row_num)

            cpu_attribute = cell_map["CPU attribute"]["data"]

            workload['wl_list'][workload_index]["cpu_attribute"] = cpu_attribute

            workload['wl_list'][workload_index]["cpu_model"] = cell_map["CPU model"]["data"] \
                if cell_map["CPU model"]["data"] else 'Intel E5-2630 v4'

            workload['wl_list'][workload_index][cpu_attribute] = \
                int(cell_map["CPU"]["data"]) if cell_map["CPU"]["data"] else 1

            workload['wl_list'][workload_index]["ram_size"] = \
                float(cell_map["Ram"]["data"]) if cell_map["Ram"]["data"] else 1

            workload['wl_list'][workload_index]["ram_opratio"] = \
                int(cell_map["Ram OP"]["data"]) if cell_map["Ram OP"]["data"] else 1

            workload['wl_list'][workload_index]["ram_size_unit"] = cell_map["Ram Unit"]["data"]

            workload['wl_list'][workload_index]['hdd_size'] = \
                float(cell_map["Disk capacity"]["data"]) if cell_map["Disk capacity"]["data"] else 1

            workload['wl_list'][workload_index]['hdd_size_unit'] = cell_map["Disk capacity Unit"]["data"]

            workload['wl_list'][workload_index]["overhead_percentage"] = cell_map["Future growth"]["data"]

            workload['wl_list'][workload_index]['replication_factor'] = int(cell_map['RF']['data'])

            workload['wl_list'][workload_index]['fault_tolerance'] = int(cell_map['FT']["data"])

            if not workload_index:
                status = edit_settings(workload, cell_map)
            if not status:
                construct_result(status, 'Wrong inputs have been given.', None, cell_map, element[0], sheet,
                                 workload['wl_list'])
                break

        try:
            constraint = cell_map['Constraint']['data']
            if constraint:
                if constraint == "CPU":
                    key = cpu_attribute
                    end_range = 5000
                elif constraint == 'Ram':
                    key = "ram_size"
                    end_range = 188416
                else:
                    key = 'hdd_size'
                    end_range = 3000000
                start_range = 1

                workload['wl_list'][workload_index][key] = start_range
                status, response, _ = get_response(workload, None, None, True)
                if status:
                    status, response, max_value = binary_search(key, start_range, end_range, workload)
                    constraint_cell = sheet.cell(row=row_num, column=cell_map[constraint]['column'])

                    if cell_map[constraint]['data'] and cell_map[constraint]['data'] != str(max_value):
                        constraint_cell.value = str(cell_map[constraint]['data']) + ' -> ' + str(max_value)
                    else:
                        constraint_cell.value = str(max_value)
                else:
                    construct_result(status, response, None, cell_map, element[0], sheet, workload['wl_list'])
                    continue
            else:
                status, response, _ = get_response(workload, None, None, True)
        except ValueError:
            status, response, max_value = False, 'Server Error. Please check this workload set', None

        construct_result(status, response, None, cell_map, element[0], sheet, workload['wl_list'])


def handle_file_upload(file_name, row_num, cell_map, sheet):

    file_dict = {'file': open(os.path.dirname(os.path.realpath(__file__)) + '/' + file_name, 'rb')}

    values = {'DB': 'photcat', 'OUT': 'csv', 'SHORT': 'short'}

    response = requests.post(file_url, data=values, headers=header, files=file_dict).json()

    if response['status'] == 'error':
        construct_result(False, response['errorMessage'], None, cell_map, row_num, sheet, None)


def test_mixed(sheet, workload, cell_map, workload_set_map):

    for element in workload_set_map:

        workload['wl_list'] = list()

        if sheet.cell(row=element[0], column=cell_map['Workload type']['column']).value == 'FILE-UPLOAD':
            handle_file_upload(sheet.cell(row=element[0], column=cell_map['File name']['column']).value, element[0],
                               cell_map, sheet)
            continue

        for row_num in range(element[0], element[1] + 1):

            workload_index = row_num - element[0]

            for key, value_dict in cell_map.items():
                if key in RESULT_SET and row_num != element[0]:
                    continue
                value_dict["data"] = sheet.cell(row=row_num, column=value_dict["column"]).value

            new_workload = deepcopy(WL_TYPE_MAPPER[cell_map["Workload type"]["data"]]['wl_list'][0])

            workload['wl_list'].append(deepcopy(new_workload))

            workload['wl_list'][workload_index]['wl_name'] = sheet.title + ' - ' + str(row_num)

            workload['wl_list'][workload_index]['replication_factor'] = int(cell_map['RF']['data'])

            workload['wl_list'][workload_index]['fault_tolerance'] = int(cell_map['FT']["data"])

            if workload['wl_list'][workload_index]['wl_type'] == 'VSI':

                workload['wl_list'][workload_index]['num_vms'] = int(cell_map['#VMs [VSI]']['data'])

                workload['wl_list'][workload_index]['profile_type'] = cell_map['Type']['data']

                default_config_type = "VSI/ROBO"

            elif workload['wl_list'][workload_index]['wl_type'] == 'VDI_INFRA':

                workload['wl_list'][workload_index]['vm_details']['vm1']['num_vms'] = \
                    int(cell_map['#VMs - 1 [INFRA]']['data'])

                workload['wl_list'][workload_index]['vm_details']['vm2']['num_vms'] = \
                    int(cell_map['#VMs - 2 [INFRA]']['data'])

                default_config_type = 'INFRA'

                workload['wl_list'][workload_index]['infra_type'] = CONVERSION_DICT[cell_map['Type']['data']]

            elif workload['wl_list'][workload_index]['wl_type'] == 'ROBO':

                workload['wl_list'][workload_index]['num_vms'] = int(cell_map['#VMs [ROBO]']['data'])

                workload['wl_list'][workload_index]['profile_type'] = cell_map['Type']['data']

                default_config_type = "VSI/ROBO"

            elif workload['wl_list'][workload_index]['wl_type'] == 'RDSH':

                workload['wl_list'][workload_index]['total_users'] = int(cell_map['#Users [RDSH]']['data'])

                workload['wl_list'][workload_index]['broker_type'] = CONVERSION_DICT[cell_map['Broker [RDSH]']['data']]

                workload['wl_list'][workload_index]['profile_type'] = CONVERSION_DICT[cell_map['Type']['data']]

                default_config_type = "RDSH"

            elif workload['wl_list'][workload_index]['wl_type'] == 'VDI':

                workload['wl_list'][workload_index]['num_desktops'] = int(cell_map['#DESKTOPs [VDI]']['data'])

                default_config_type = "VDI"

                workload['wl_list'][workload_index]['os_type'] = CONVERSION_DICT[cell_map['OS [VDI]']['data']]

                workload['wl_list'][workload_index]['profile_type'] = CONVERSION_DICT[cell_map['Type']['data']]

                workload['wl_list'][workload_index]['provisioning_type'] = \
                    CONVERSION_DICT[cell_map['Provisioning [VDI]']["data"]]

                if 'Home Directory [VDI]' in cell_map and cell_map['Home Directory [VDI]']['data'] == 'yes':
                    workload['wl_list'][workload_index]['vdi_directory'] = True
                else:
                    workload['wl_list'][workload_index]['vdi_directory'] = False

                if cell_map["GPU [VDI]"]["data"] == "Enabled":

                    workload['wl_list'][workload_index]["gpu_users"] = 1

                    if "video_RAM [VDI]" in cell_map and cell_map["video_RAM [VDI]"]["data"]:
                        workload['wl_list'][workload_index]["video_RAM"] = int(cell_map["video_RAM [VDI]"]["data"])
                    else:
                        workload['wl_list'][workload_index]["video_RAM"] = 1024

                else:
                    workload['wl_list'][workload_index]["gpu_users"] = 0

            elif workload['wl_list'][workload_index]['wl_type'] in ['DB', 'ORACLE']:

                workload['wl_list'][workload_index]['num_db_instances'] = int(cell_map['#DBs [DB]']['data'])

                workload['wl_list'][workload_index]['profile_type'] = cell_map['Type']['data']

                workload['wl_list'][workload_index]['db_type'] = cell_map['DB type [DB]']['data']

                if workload['wl_list'][workload_index]['wl_type'] == 'ORACLE':

                    workload['wl_list'][workload_index]['internal_type'] = \
                        "O" + workload['wl_list'][workload_index]['db_type']

                    default_config_type = 'ORACLE'

                else:

                    workload['wl_list'][workload_index]['internal_type'] = \
                        workload['wl_list'][workload_index]['db_type']

                    default_config_type = 'SQL'

            else:
                default_config_type = None

            if not workload_index:
                status = edit_settings(workload, cell_map)
            if not status:
                construct_result(status, 'Wrong inputs have been given.', None, cell_map, element[0], sheet,
                                 workload['wl_list'])
                break

            # assumed that only default type profiles are present in template
            if workload['wl_list'][workload_index]['wl_type'] in ['VDI_INFRA', 'VEEAM', 'SPLUNK']:
                profile_type = None
            else:
                profile_type = workload['wl_list'][workload_index]['profile_type']

            configure_default_profile(workload, default_config_type, profile_type, workload_index)

        try:
            status, response, max_value = get_response(workload, None, None, True)
        except ValueError:
            status, response, max_value = False, 'Server Error. Please check this workload set', None

        construct_result(status, response, max_value, cell_map, element[0], sheet, workload['wl_list'])


def remove_transitions():

    for sheet_name in workload_tests:

        sheet = book.get_sheet_by_name(sheet_name)
        cell_map = get_cell_map(sheet)

        for cell_key in cell_map:

            for row in range(2, sheet.max_row + 1):

                cell = sheet.cell(row=row, column=cell_map[cell_key]['column'])

                if cell.value:
                    if ' -> ' in str(cell.value):
                        value_to_be = str(cell.value.split(' -> ')[1])
                    else:
                        value_to_be = str(cell.value)

                    if value_to_be.isdigit():
                        value_to_be = int(value_to_be)

                    cell.value = value_to_be

                cell.fill = whitefill


def tester(sheet_name, sheet_reference):

    cell_map = get_cell_map(sheet_reference)

    workload_set_map = get_row_map(sheet_reference, cell_map["Workload Set"]["column"])

    error_count[sheet_name]['tests'] = len(workload_set_map)

    eval(workload_import_mapper[sheet_name]['function'])(sheet_reference,
                                                         workload_import_mapper[sheet_name]['workload'],
                                                         cell_map, workload_set_map)


# Install: pip install openpyxl==2.6.4
dir_name = os.path.dirname(os.path.realpath(__file__))
book = openpyxl.load_workbook(dir_name + '/' + TEMPLATE)
workload_tests = book.get_sheet_names()

host_url = "http://127.0.0.1:8000/hyperconverged/scenarios/%s" % SCENARIO_ID
file_url = "http://127.0.0.1:8000/hyperconverged/processesxstat"
header = {'Authorization': '03011f56614a48b7e76642149b9e1e7ecad55fa2'}
version = requests.get('http://127.0.0.1:8000/hyperconverged/version', headers=header).json()["sizer_version"]

yellowfill = PatternFill(start_color=colors.YELLOW, end_color=colors.YELLOW, fill_type='solid')
redfill = PatternFill(start_color=colors.RED, end_color=colors.RED, fill_type='solid')
whitefill = PatternFill(start_color='00000000', end_color='FFFFFFFF', fill_type=None)

workload_import_mapper = defaultdict(dict)
for sheet_name in workload_tests:
    if 'VDI' in sheet_name:
        handler_function = 'test_vdi'
        wl = VDI
    elif 'VSI' in sheet_name:
        handler_function = 'test_vsi_robo'
        wl = VSI
    elif 'ROBO' in sheet_name:
        handler_function = 'test_vsi_robo'
        wl = ROBO
    elif 'DB' in sheet_name:
        handler_function = 'test_db'
        wl = DB
    elif 'ORACLE' in sheet_name:
        handler_function = 'test_db'
        wl = ORACLE
    elif 'RAW' in sheet_name:
        handler_function = 'test_raw'
        wl = RAW
    elif 'INFRA' in sheet_name:
        handler_function = 'test_infra'
        wl = VDI_INFRA
    elif 'VEEAM' in sheet_name:
        handler_function = 'test_veeam'
        wl = VEEAM
    elif 'SPLUNK' in sheet_name:
        handler_function = 'test_splunk'
        wl = SPLUNK
    elif 'RDSH' in sheet_name:
        handler_function = 'test_rdsh'
        wl = RDSH
    elif 'AIML' in sheet_name:
        handler_function = 'test_aiml'
        wl = AIML
    else:
        handler_function = 'test_mixed'
        wl = MIXED

    workload_import_mapper[sheet_name]['workload'] = wl
    workload_import_mapper[sheet_name]['function'] = handler_function

error_count = dict()

destination_path = dir_name + '/' + RESULT_FILE_NAME
body = ''
# email_list = ["srivathsa.mugalodi@maplelabs.com", "vinayak.bhakta@maplelabs.com", "savitha.ramappa@maplelabs.com",
#               "michael.wang@maplelabs.com", "thouqueer.ahmed@maplelabs.com", "hm.chandana@Maplelabs.com"]

email_list = ["michael.wang@maplelabs.com", "savitha.ramappa@maplelabs.com"]

message = Message(From="maple-noreply@maplelabs.com",
                  To=email_list,
                  BCC=[],
                  Subject="Hyperflex-Sizer Test Results w.r.t " + argv[1],
                  charset='utf-8')

initialise_test_dict()

for sheet_name in workload_tests:
    print (sheet_name)
    tester(sheet_name, book.get_sheet_by_name(sheet_name))

error_num = 0
book.save(destination_path)
if argv[1] == 'Master' and TEMPLATE == 'Master.xlsx':
    remove_transitions()
    book.save(dir_name + '/' + TEMPLATE)

for sheet_name in error_count:
    error_count[sheet_name]["success"] = error_count[sheet_name]["tests"] - (error_count[sheet_name]["regressions"] +
                                                                             error_count[sheet_name]["expected errors"])
    body += str(sheet_name) + '\n'
    for key in error_count[sheet_name]:
        body += '\t' + str(key) + ': ' + str(error_count[sheet_name][key]) + '\n'

message.Body = body
message.attach(destination_path)
sender = Mailer('webmail.xoriant.com')
sender.send(message)
